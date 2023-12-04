import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
from PyQt5.QtGui import QColor
from openpyxl import load_workbook

class ExcelTableWidget(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        self.table = QTableWidget(self)
        self.central_layout = QVBoxLayout(self.central_widget)
        self.central_layout.addWidget(self.table)

        self.load_excel_data("rutput_data (1).xlsx")

        self.setGeometry(100, 100, 800, 600)
        self.setWindowTitle('Excel Table Viewer')
        self.show()

    def load_excel_data(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                current_row = self.table.rowCount()
                self.table.insertRow(current_row)
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(current_row, col, item)

                   
                    if col != 0:
                        
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

        except Exception as e:
            print(f"Error loading Excel data: {e}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelTableWidget()
    sys.exit(app.exec_())
