import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QLabel
from PyQt5.QtGui import QColor
from openpyxl import load_workbook


class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1920, 1080)
        Form.setStyleSheet("#leftbox {\n"
"    background-image: url(\'C:/coding/bg.jpg\');\n"
"    background-size: 80%;\n"
"}\n"
"\n"
"#mainbody {\n"
"    background-color: #ffffff;\n"
"}\n"
"\n"
"#card1, #card2 {\n"
"    border: 1px solid black;\n"
"    border-radius: 20px;\n"
"}\n"
"\n"
"QPushButton {\n"
"    background-color: transparent;\n"
"    color: #555555;\n"
"    font-size: 40px;\n"
"}\n"
"QPushButton:hover {\n"
"  background-color: rgba(0, 0, 0, 0.1);\n"
"}\n"
"\n"
"QComboBox{\n"
"        background-color: transparent;\n"
"        color: white;\n"
"    font-size: 20px;\n"
"}\n"
"\n"
"#graphicsView{\n"
"    border: 1px solid black;\n"
"}")


        self.mainbody = QtWidgets.QWidget(Form)
        self.mainbody.setGeometry(QtCore.QRect(0, 0, 1920, 1080))
        self.mainbody.setObjectName("mainbody")

        self.leftbox = QtWidgets.QWidget(self.mainbody)
        self.leftbox.setGeometry(QtCore.QRect(0, 0, 300, 1080))
        self.leftbox.setObjectName("leftbox")

        self.pushButton = QtWidgets.QPushButton(self.leftbox)
        self.pushButton.setGeometry(QtCore.QRect(0, 50, 300, 51))
        self.pushButton.setObjectName("pushButton")

        self.comboBox = QtWidgets.QComboBox(self.leftbox)
        self.comboBox.setGeometry(QtCore.QRect(0, 140, 300, 41))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")

        self.comboBox_2 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_2.setGeometry(QtCore.QRect(0, 220, 300, 41))
        self.comboBox_2.setObjectName("comboBox_2")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")

        self.comboBox_3 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_3.setGeometry(QtCore.QRect(0, 300, 300, 31))
        self.comboBox_3.setObjectName("comboBox_3")
        self.comboBox_3.addItem("")

        self.comboBox_4 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_4.setGeometry(QtCore.QRect(0, 380, 300, 31))
        self.comboBox_4.setObjectName("comboBox_4")
        self.comboBox_4.addItem("")

        self.comboBox_5 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_5.setGeometry(QtCore.QRect(0, 460, 300, 41))
        self.comboBox_5.setObjectName("comboBox_5")
        self.comboBox_5.addItem("")

        self.comboBox_6 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_6.setGeometry(QtCore.QRect(0, 540, 300, 41))
        self.comboBox_6.setObjectName("comboBox_6")
        self.comboBox_6.addItem("")

        self.comboBox_7 = QtWidgets.QComboBox(self.leftbox)
        self.comboBox_7.setGeometry(QtCore.QRect(0, 620, 300, 41))
        self.comboBox_7.setObjectName("comboBox_7")
        self.comboBox_7.addItem("")

        self.card1 = QtWidgets.QWidget(self.mainbody)
        self.card1.setGeometry(QtCore.QRect(550, 30, 400, 450))
        self.card1.setObjectName("card1")

        # Create a QLabel for displaying text
        self.card1_label = QLabel(self.card1)
        self.card1_label.setGeometry(QtCore.QRect(0, 0, 400, 450))
        self.card1_label.setObjectName("card1_label")

        # Set stylesheets for center alignment
        self.card1_label.setStyleSheet("QLabel { text-align: center; }")

    

        self.card2 = QtWidgets.QWidget(self.mainbody)
        self.card2.setGeometry(QtCore.QRect(550, 500, 400, 450))
        self.card2.setObjectName("card2")

        self.tableView = QtWidgets.QTableWidget(self.mainbody)
        self.tableView.setGeometry(QtCore.QRect(1000, 30, 800, 450))
        self.tableView.setObjectName("tableView")

        self.graphicsView = QtWidgets.QGraphicsView(self.mainbody)
        self.graphicsView.setGeometry(QtCore.QRect(1000, 500, 800, 450))
        self.graphicsView.setObjectName("graphicsView")



        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.pushButton.setText(_translate("Form", "Dashboard"))
        
        self.comboBox.setItemText(0, _translate("Form", "Gentral Info A"))
        self.comboBox.setItemText(1, _translate("Form", "Gentral Info B"))
        self.comboBox.setItemText(2, _translate("Form", "Gentral Info C"))
    
    
        self.comboBox_2.setItemText(0, _translate("Form", "Annual"))
        self.comboBox_2.setItemText(1, _translate("Form", "New Item"))
        self.comboBox_2.setItemText(2, _translate("Form", "New Item"))
        self.comboBox_3.setItemText(0, _translate("Form", "Actual "))
        self.comboBox_4.setItemText(0, _translate("Form", "Gap Analysis Recurring Revenue and Recurring Expenses"))
        self.comboBox_5.setItemText(0, _translate("Form", "Total  Fund Allocation across Programs"))
        self.comboBox_6.setItemText(0, _translate("Form", "The way Revenue  and Allocations Planned "))
        self.comboBox_7.setItemText(0, _translate("Form", " Own Revenue Planned Per citizen"))
        








    def loadExcelData(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col != 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

       

            # Calculate and display row sum in Card 1
            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")

            
    def displayRowSum(self, sheet):
        # Display the sum of each row in Card 1
        row_sums = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_sum = sum(value for value in row if isinstance(value, (int, float)))
            row_sums.append(row_sum)

        # Set text in QLabel
        self.card1_label.setText("VExhiles: " + ', '.join(map(str, row_sums)))



class MyForm(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.loadExcelData("rutput_data (1).xlsx")

        # Connect the currentIndexChanged signal of the comboBox to the updateMainBox method
        self.comboBox.currentIndexChanged.connect(self.updateMainBox)

    def updateMainBox(self, index):
        # You can use the index to determine which item is selected
        selected_item = self.comboBox.itemText(index)

        # Update the main box content based on the selected item
        if selected_item == "Gentral Info A":
            self.updateMainBoxForInfoA()
        elif selected_item == "Gentral Info B":
            self.updateMainBoxForInfoB()
        elif selected_item == "Gentral Info C":
            self.updateMainBoxForInfoC()
        # Add more conditions as needed for other items

    def updateMainBoxForInfoA(self):
        try:
            file_path = "rutput_data (1).xlsx"  # Replace with the actual file path
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col != 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

            # Calculate and display row sum in Card 1
            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")


    def updateMainBoxForInfoB(self):
        try:
            file_path = "rutput_data (1).xlsx"  # Replace with the actual file path
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col == 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

            # Calculate and display row sum in Card 1
            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")

        
        

    def updateMainBoxForInfoC(self):
        try:
            file_path = "rutput_data (1).xlsx"  # Replace with the actual file path
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col != 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

            # Calculate and display row sum in Card 1
            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyForm()
    window.show()
    sys.exit(app.exec_())
